import os
import re
import io
import openpyxl
from openpyxl.styles import PatternFill, Font
from telegram import Update
from telegram.ext import ContextTypes, CommandHandler
from telegram.constants import ParseMode
from database.db import get_db

OWNER_ID = int(os.getenv("OWNER_ID", "0"))
TOTAL_STUDENTS = 59  # 📌 ကျောင်းသား စုစုပေါင်း အရေအတွက်

# ================= 📋 ကျောင်းသား စာရင်းသွင်းမှု စစ်ဆေးခြင်း (/list) =================
async def list_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != OWNER_ID: return
    
    db = get_db()
    students = await db.student_info.find({}).to_list(length=None)
    
    registered_count = len(students)
    missing_count = TOTAL_STUDENTS - registered_count
    
    # ရထားပြီးသား ခုံနံပါတ်များမှ ဂဏန်းများကို ဆွဲထုတ်ခြင်း (ဥပမာ "Sem-II-01" မှ "1" ကို ယူမည်)
    registered_nums = set()
    for s in students:
        roll = s.get("roll_no", "")
        match = re.search(r'(\d+)$', roll)
        if match:
            registered_nums.add(int(match.group(1)))
            
    # ကျန်နေသေးသော ဂဏန်းများကို ရှာခြင်း
    missing_nums = []
    for i in range(1, TOTAL_STUDENTS + 1):
        if i not in registered_nums:
            missing_nums.append(f"{i:02d}") # "01, 02" ပုံစံပြရန်
            
    missing_str = ", ".join(missing_nums) if missing_nums else "🎉 အားလုံး စာရင်းသွင်းပြီးပါပြီ!"

    msg = (
        f"📊 <b>ကျောင်းသား စာရင်းသွင်းမှု အခြေအနေ</b>\n"
        f"━━━━━━━━━━━━━━━━━━\n"
        f"🔹 စာရင်းသွင်းပြီးသူ: <b>{registered_count}</b> ယောက်\n"
        f"🔹 စာရင်းသွင်းရန် ကျန်ရှိသူ: <b>{missing_count}</b> ယောက်\n\n"
        f"⏳ <b>စာရင်းသွင်းရန် ကျန်သေးသည့် ခုံနံပါတ်များ:</b>\n"
        f"<code>{missing_str}</code>"
    )
    await update.message.reply_text(msg, parse_mode=ParseMode.HTML)


# ================= 📊 ရန်ပုံငွေ စာရင်းချုပ်စစ်ဆေးခြင်း (/report) =================
async def report_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != OWNER_ID: return
    
    if len(context.args) < 1:
        await update.message.reply_text("⚠️ အသုံးပြုနည်း: <code>/report &lt;FundID&gt;</code> (ဥပမာ: /report sports)", parse_mode=ParseMode.HTML)
        return
        
    fund_id = context.args[0].lower()
    db = get_db()
    
    fund = await db.funds.find_one({"fund_id": fund_id})
    if not fund:
        await update.message.reply_text(f"⚠️ <code>{fund_id}</code> အမည်ဖြင့် ရန်ပုံငွေ မရှိပါ။", parse_mode=ParseMode.HTML)
        return
        
    target = fund.get("target_amount", 0)
    students = await db.student_info.find({}).sort("roll_no", 1).to_list(length=None)
    
    full, partial, unpaid = [], [], []
    
    for s in students:
        paid = s.get("paid_funds", {}).get(fund_id, 0)
        roll_clean = s.get("roll_no").split('-')[-1] # နောက်ဆုံးဂဏန်းပဲ ယူမည်
        name_str = f"<code>{roll_clean}</code> - {s.get('name')}"
        
        if paid >= target:
            full.append(name_str)
        elif paid > 0:
            partial.append(f"{name_str} ({paid}/{target})")
        else:
            unpaid.append(name_str)
            
    # Message တည်ဆောက်ခြင်း
    msg = f"📋 <b>REPORT: {fund.get('fund_name')}</b>\n━━━━━━━━━━━━━━━━━━\n\n"
    
    msg += f"✅ <b>အပြည့်ပေးပြီးသူများ ({len(full)}):</b>\n"
    msg += "\n".join(full) if full else "➖ မရှိသေးပါ"
    msg += "\n\n"
    
    msg += f"⚠️ <b>တစ်စိတ်တစ်ပိုင်း ပေးထားသူများ ({len(partial)}):</b>\n"
    msg += "\n".join(partial) if partial else "➖ မရှိသေးပါ"
    msg += "\n\n"
    
    msg += f"❌ <b>လုံးဝ မပေးရသေးသူများ ({len(unpaid)}):</b>\n"
    msg += "\n".join(unpaid) if unpaid else "➖ မရှိသေးပါ"

    await update.message.reply_text(msg, parse_mode=ParseMode.HTML)


# ================= 💰 ဘဏ္ဍာငွေ အခြေအနေ ခြုံငုံကြည့်ရန် (/vault) =================
async def vault_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != OWNER_ID: return
    
    db = get_db()
    funds = await db.funds.find({}).to_list(length=None)
    if not funds:
        return await update.message.reply_text("⚠️ ကောက်ခံနေသော ရန်ပုံငွေ မရှိသေးပါ။")
        
    students = await db.student_info.find({}).to_list(length=None)
    
    msg = "💰 <b>CLASS VAULT DASHBOARD</b>\n━━━━━━━━━━━━━━━━━━\n\n"
    
    for f in funds:
        fid = f.get("fund_id")
        target = f.get("target_amount", 0)
        total_expected = target * TOTAL_STUDENTS
        
        total_collected = sum(s.get("paid_funds", {}).get(fid, 0) for s in students)
        
        msg += f"🎯 <b>{f.get('fund_name')}</b>\n"
        msg += f"🔹 ရရှိပြီးငွေ: <code>{total_collected:,}</code> MMK\n"
        msg += f"🔸 ခန့်မှန်း စုစုပေါင်း: <code>{total_expected:,}</code> MMK\n\n"
        
    await update.message.reply_text(msg, parse_mode=ParseMode.HTML)


# ================= 🏠 အဆောင်နေ ကျောင်းသားများ စာရင်း (/hostel) =================
async def hostel_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != OWNER_ID: return
    
    db = get_db()
    students = await db.student_info.find({"hostel": True}).sort("roll_no", 1).to_list(length=None)
    
    if not students:
        return await update.message.reply_text("⚠️ အဆောင်နေ ကျောင်းသား မရှိသေးပါ။")
        
    msg = f"🏠 <b>အဆောင်နေ ကျောင်းသားများ စာရင်း ({len(students)} ယောက်)</b>\n━━━━━━━━━━━━━━━━━━\n\n"
    
    for s in students:
        roll_clean = s.get("roll_no").split('-')[-1]
        phone = s.get("phone", "➖")
        tg_username = f" (@{s.get('tg_username')})" if s.get("tg_username") else ""
        
        msg += f"🔹 <code>{roll_clean}</code> - {s.get('name')}{tg_username}\n    📞 {phone}\n\n"
        
    await update.message.reply_text(msg, parse_mode=ParseMode.HTML)

# ================= 📂 စာရင်းဇယားများကို CSV ဖိုင်ဖြင့် ထုတ်ယူခြင်း (/export) =================
async def export_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != OWNER_ID: return
    
    db = get_db()
    # Database မှ Data အကုန်အရင်ဆွဲထုတ်မည်
    students = await db.student_info.find({}).to_list(length=None)
    funds = await db.funds.find({}).to_list(length=None)
    
    if not students:
        return await update.message.reply_text("⚠️ Database ထဲတွင် ကျောင်းသား အချက်အလက် မရှိသေးပါ။")
        
    # 📌 Roll No ၏ နောက်ဆုံးဂဏန်းကို အခြေခံ၍ အတိအကျ အစဉ်လိုက်စီခြင်း (Sorting)
    def get_roll_num(student):
        roll = student.get("roll_no", "")
        match = re.search(r'(\d+)$', roll)
        return int(match.group(1)) if match else 9999  # ဂဏန်းမပါလျှင် အောက်ဆုံးပို့မည်

    students.sort(key=get_roll_num)
    
    # ရန်ပုံငွေ ခေါင်းစဉ်များကို ဆွဲထုတ်ခြင်း
    fund_ids = [f["fund_id"] for f in funds if "fund_id" in f]
    fund_names = [f.get("fund_name", f["fund_id"]) for f in funds if "fund_id" in f]
    
    # မှတ်ဉာဏ် (Memory) ပေါ်တွင် CSV ဖိုင် တည်ဆောက်ခြင်း
    output = io.StringIO()
    writer = csv.writer(output)
    
    # 📌 Student ID ကို ဖယ်ရှားပြီး ခေါင်းစဉ်တန်း (Header Row) ထည့်ခြင်း
    header = ["Roll No", "Name", "Phone", "Telegram Username", "Hostel"] + fund_names
    writer.writerow(header)
    
    # ကျောင်းသား တစ်ယောက်ချင်းစီ၏ Data များကို ထည့်ခြင်း
    for s in students:
        row = [
            s.get("roll_no", "➖"),
            s.get("name", "➖"),
            s.get("phone", "➖"),
            s.get("tg_username", "➖"),
            "Yes" if s.get("hostel") else "No"
        ]
        # ရန်ပုံငွေ အလိုက် သွင်းထားငွေကို Column အလိုက် ထည့်ခြင်း
        for fid in fund_ids:
            row.append(str(s.get("paid_funds", {}).get(fid, 0)))
            
        writer.writerow(row)
        
    # ဖိုင်ကို Telegram သို့ ပို့ဆောင်ခြင်း
    output.seek(0)
    await update.message.reply_document(
        document=output.getvalue().encode('utf-8-sig'), # utf-8-sig ကြောင့် Excel တွင် မြန်မာစာ မှန်ကန်မည်
        filename="Class_Report.csv",
        caption="✅ <b>Database Export အောင်မြင်ပါသည်။</b>\n\nဒီဖိုင်ကို Excel သို့မဟုတ် Google Sheets ဖြင့် ဖွင့်ကြည့်နိုင်ပါသည်။",
        parse_mode=ParseMode.HTML
    )

# ================= 📂 သတ်မှတ် Fund ID အလိုက် Excel (.xlsx) ထုတ်ယူခြင်း (/exportfund) =================
async def export_fund_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != OWNER_ID: return
    
    if len(context.args) < 1:
        await update.message.reply_text(
            "⚠️ <b>အသုံးပြုနည်း:</b> <code>/exportfund &lt;FundID&gt;</code>\n"
            "(ဥပမာ - <code>/exportfund sports</code>)", 
            parse_mode=ParseMode.HTML
        )
        return
        
    fund_id = context.args[0].lower()
    db = get_db()
    
    fund = await db.funds.find_one({"fund_id": fund_id})
    if not fund:
        await update.message.reply_text(f"⚠️ <code>{fund_id}</code> အမည်ဖြင့် ရန်ပုံငွေ မရှိပါ။", parse_mode=ParseMode.HTML)
        return
        
    fund_name = fund.get("fund_name", fund_id)
    students = await db.student_info.find({}).to_list(length=None)
    
    if not students:
        return await update.message.reply_text("⚠️ Database ထဲတွင် ကျောင်းသား အချက်အလက် မရှိသေးပါ။")

    export_data = []
    total_collected = 0
    
    for s in students:
        roll = s.get("roll_no", "➖")
        match = re.search(r'(\d+)$', roll)
        roll_num = int(match.group(1)) if match else 9999
        
        name = s.get("name", "➖")
        paid = s.get("paid_funds", {}).get(fund_id, 0)
        
        # ပိုက်ဆံသွင်းထားလျှင် Paid၊ မသွင်းထားလျှင် Unpaid သတ်မှတ်မည်
        status = "Paid" if paid > 0 else "Unpaid"
            
        export_data.append({
            "roll": roll,
            "roll_num": roll_num,
            "name": name,
            "status": status,
            "paid": paid
        })
        total_collected += paid
        
    # 📌 ငွေသွင်းသည့် ပမာဏအလိုက် Sorting စီခြင်း (ငွေများသူ အပေါ်ဆုံး၊ ငွေတူပါက ခုံနံပါတ်စဉ်အလိုက်)
    export_data.sort(key=lambda x: (-x["paid"], x["roll_num"]))
    
    # ================= EXCEL ဖန်တီးခြင်း =================
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fund Report"
    
    # 📌 ခေါင်းစဉ်တန်း (Headers)
    headers = ["Roll No", "Name", "Status", f"Paid Amount ({fund_name})"]
    ws.append(headers)
    for col in range(1, 5):
        ws.cell(row=1, column=col).font = Font(bold=True)
        
    # 📌 အရောင် (Colors) သတ်မှတ်ခြင်း (Professional Excel Colors)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(color="006100")
    
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006")
    
    # 📌 Data များထည့်သွင်းခြင်း
    for row_idx, d in enumerate(export_data, start=2):
        ws.cell(row=row_idx, column=1, value=d["roll"])
        ws.cell(row=row_idx, column=2, value=d["name"])
        
        # Status Column အတွက် အရောင်ချယ်ခြင်း
        status_cell = ws.cell(row=row_idx, column=3, value=d["status"])
        if d["status"] == "Paid":
            status_cell.fill = green_fill
            status_cell.font = green_font
        else:
            status_cell.fill = red_fill
            status_cell.font = red_font
            
        # ငွေပမာဏ Column ထည့်ခြင်း
        ws.cell(row=row_idx, column=4, value=d["paid"])
        
    # 📌 အောက်ဆုံးတွင် TOTAL စုစုပေါင်း ထည့်ခြင်း
    last_row = len(export_data) + 3
    ws.cell(row=last_row, column=3, value="TOTAL:").font = Font(bold=True)
    ws.cell(row=last_row, column=4, value=total_collected).font = Font(bold=True)
    
    # 📌 ဖိုင်ကို Memory ပေါ်တွင် သိမ်းပြီး Telegram သို့ ပို့ဆောင်ခြင်း
    import io
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    await update.message.reply_document(
        document=output,
        filename=f"{fund_name}_Report.xlsx",
        caption=(
            f"✅ <b>{fund_name} အတွက် စာရင်းထုတ်ပြီးပါပြီ။</b>\n\n"
            f"🔹 စုစုပေါင်း ရငွေ: <b>{total_collected:,} MMK</b>\n"
            f"<i>(ငွေသွင်းထားမှု အများဆုံးမှ အနည်းဆုံးသို့ အစဉ်လိုက် စီပေးထားပါသည်။)</i>"
        ),
        parse_mode=ParseMode.HTML
    )